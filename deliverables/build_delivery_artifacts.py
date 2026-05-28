from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Flowable,
    Image,
    KeepTogether,
    ListFlowable,
    ListItem,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ROOT = Path(__file__).resolve().parent
PDF_PATH = ROOT / "AI智能客服售前跟进系统_技术方案手册_v1.0.pdf"
XLSX_PATH = ROOT / "AI智能客服售前跟进系统_工程报价单_v1.0.xlsx"


PROJECT_NAME = "AI智能客服售前跟进系统"
VERSION = "v1.0"
DOC_DATE = "2026-05-20"
UNIT_PRICE = 1500


QUOTE_ROWS = [
    ("1. 线索与销售任务", "手机号线索导入/录入", 0.25, "支持导入小风车手机号线索"),
    ("1. 线索与销售任务", "线索状态流转", 0.25, "记录待添加、已申请、已通过、已接管等状态"),
    ("1. 线索与销售任务", "销售库/微信账号配置", 0.25, "配置销售及对应微信设备"),
    ("1. 线索与销售任务", "简单分配规则", 0.50, "按销售容量或手动规则分配线索"),
    ("1. 线索与销售任务", "加好友任务生成", 0.50, "根据线索生成微信添加任务"),
    ("1. 线索与销售任务", "任务列表/基础筛选", 0.25, "查看任务状态和失败任务"),
    ("2. RPA加好友与备注绑定", "手机号搜索流程", 1.00, "控制微信按手机号搜索客户"),
    ("2. RPA加好友与备注绑定", "固定好友申请语", 0.50, "发送统一添加好友话术"),
    ("2. RPA加好友与备注绑定", "提交申请与结果记录", 0.50, "记录申请成功、失败、异常"),
    ("2. RPA加好友与备注绑定", "备注规则生成", 0.50, "生成线索短码备注"),
    ("2. RPA加好友与备注绑定", "会话标题/备注匹配绑定", 0.50, "将微信会话绑定回线索"),
    ("2. RPA加好友与备注绑定", "RPA异常重试与失败原因", 0.50, "处理搜索失败、控件异常、发送失败"),
    ("2. RPA加好友与备注绑定", "单机Worker心跳/状态上报", 0.50, "上报销售电脑在线和运行状态"),
    ("3. 文字AI回复链路", "接入现有OmniAuto文字回复", 2.50, "复用已有微信AI客服文字能力"),
    ("3. 文字AI回复链路", "会话上下文绑定线索", 2.00, "让AI知道当前客户对应哪条线索"),
    ("3. 文字AI回复链路", "最多轮次/停止规则", 1.50, "控制最多回复轮次和停止条件"),
    ("3. 文字AI回复链路", "销售手动回复后AI停止", 2.00, "检测销售发言后暂停AI"),
    ("3. 文字AI回复链路", "高风险问题转人工", 2.00, "价格、车况、金融等敏感问题转人工"),
    ("3. 文字AI回复链路", "回复审计记录", 1.50, "保存AI回复、原因和命中规则"),
    ("4. 图片采集与视觉理解", "图片消息识别（type=3）", 1.50, "识别微信图片消息"),
    ("4. 图片采集与视觉理解", "RPA点开图片", 2.50, "自动打开客户发来的图片"),
    ("4. 图片采集与视觉理解", "图片另存本地", 2.00, "保存图片供模型识别"),
    ("4. 图片采集与视觉理解", "图片资产入库", 1.50, "记录图片路径、来源会话和线索"),
    ("4. 图片采集与视觉理解", "千问视觉API接入", 2.50, "调用视觉模型理解图片内容"),
    ("4. 图片采集与视觉理解", "图片理解结果结构化", 2.50, "提取车型、颜色、预算、用途等信息"),
    ("5. 图文回复训练与编排", "二手车图片咨询场景梳理", 2.00, "梳理客户发图常见意图"),
    ("5. 图文回复训练与编排", "图文回复样本构建", 2.50, "构建图文对话训练/调优样本"),
    ("5. 图文回复训练与编排", "图片理解结果接入evidence pack", 2.00, "将图片识别结果交给OmniAuto回复引擎"),
    ("5. 图文回复训练与编排", "图文回复提示词/风格规则", 2.00, "调整图文回复口吻和销售风格"),
    ("5. 图文回复训练与编排", "图文回复Guard规则", 2.00, "防止图片回复乱承诺价格、车况、金融"),
    ("5. 图文回复训练与编排", "图片回复回归测试集", 1.50, "固化图片场景测试用例"),
    ("6. 大风车与车源索引", "大风车鉴权签名封装", 0.50, "封装appKey/appSecret签名"),
    ("6. 大风车与车源索引", "店铺/车辆ID拉取适配", 0.50, "拉取指定店铺车源ID"),
    ("6. 大风车与车源索引", "车辆详情字段清洗", 0.50, "清洗品牌、车型、里程、价格等字段"),
    ("6. 大风车与车源索引", "车辆图片拉取适配", 0.50, "获取车辆图片链接"),
    ("6. 大风车与车源索引", "AI可见字段白名单", 0.50, "隔离底价、车主信息等敏感字段"),
    ("6. 大风车与车源索引", "本地车源索引", 0.50, "建立本地可检索车源数据"),
    ("7. 测试部署与交付", "Windows测试环境部署", 1.50, "部署销售电脑运行环境"),
    ("7. 测试部署与交付", "端到端流程联调", 2.50, "验证线索到接管完整流程"),
    ("7. 测试部署与交付", "失败任务/异常场景测试", 2.00, "测试加好友失败、图片失败、AI停止等场景"),
    ("7. 测试部署与交付", "真实微信灰度测试支持", 2.50, "使用真实微信账号小范围试跑"),
    ("7. 测试部署与交付", "交付文档/操作说明", 1.50, "提供部署、使用和常见问题说明"),
    ("8. 项目管理与验收协调", "需求细化与范围确认", 0.50, "确认一期正式工程范围和边界"),
    ("8. 项目管理与验收协调", "销售规则确认与待办跟进", 1.00, "跟进接管规则、大风车参数等待办"),
    ("8. 项目管理与验收协调", "周期沟通与进度同步", 0.50, "阶段性同步进度和风险"),
    ("8. 项目管理与验收协调", "验收用例整理", 0.50, "整理交付验收清单"),
    ("8. 项目管理与验收协调", "上线问题协调", 0.50, "协调试运行期间问题"),
]


def money(value: float) -> str:
    return f"¥{value:,.0f}"


def generate_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "报价明细"
    ws.freeze_panes = "A11"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subtotal_fill = PatternFill("solid", fgColor="EAF2F8")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = "AI智能客服售前跟进系统 工程报价单"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("项目名称", PROJECT_NAME),
        ("报价版本", VERSION),
        ("报价日期", DOC_DATE),
        ("人天单价", UNIT_PRICE),
        ("总工程量", ""),
        ("项目总价", ""),
    ]
    for idx, (k, v) in enumerate(meta, 3):
        ws.cell(idx, 1, k)
        ws.cell(idx, 2, v)
        ws.cell(idx, 1).font = Font(bold=True)
        ws.cell(idx, 2).fill = note_fill if idx in (6, 7) else PatternFill(fill_type=None)
    ws["B6"].number_format = '"¥"#,##0'
    ws["B7"].number_format = "0.00"
    ws["B8"].number_format = '"¥"#,##0'

    headers = ["一级模块", "子模块", "备注", "人天单价", "人天", "金额"]
    header_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    start_row = header_row + 1
    row = start_row
    current_module = None
    module_start = row
    for module, submodule, days, note in QUOTE_ROWS:
        if current_module is None:
            current_module = module
            module_start = row
        elif module != current_module:
            ws.cell(row, 1, current_module)
            ws.cell(row, 2, "小计")
            ws.cell(row, 5, f"=SUM(E{module_start}:E{row-1})")
            ws.cell(row, 6, f"=SUM(F{module_start}:F{row-1})")
            for col in range(1, 7):
                c = ws.cell(row, col)
                c.fill = subtotal_fill
                c.font = Font(bold=True)
                c.border = border
            row += 1
            current_module = module
            module_start = row

        ws.cell(row, 1, module)
        ws.cell(row, 2, submodule)
        ws.cell(row, 3, note)
        ws.cell(row, 4, UNIT_PRICE)
        ws.cell(row, 5, days)
        ws.cell(row, 6, f"=D{row}*E{row}")
        for col in range(1, 7):
            c = ws.cell(row, col)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    ws.cell(row, 1, current_module)
    ws.cell(row, 2, "小计")
    ws.cell(row, 5, f"=SUM(E{module_start}:E{row-1})")
    ws.cell(row, 6, f"=SUM(F{module_start}:F{row-1})")
    for col in range(1, 7):
        c = ws.cell(row, col)
        c.fill = subtotal_fill
        c.font = Font(bold=True)
        c.border = border
    row += 1

    ws.cell(row, 1, "合计")
    ws.cell(row, 5, f"=SUMIF(B9:B{row-1},\"小计\",E9:E{row-1})")
    ws.cell(row, 6, f"=SUMIF(B9:B{row-1},\"小计\",F9:F{row-1})")
    for col in range(1, 7):
        c = ws.cell(row, col)
        c.fill = title_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.border = border
    total_row = row
    ws["B7"] = f"=E{total_row}"
    ws["B8"] = f"=F{total_row}"

    row += 3
    ws.cell(row, 1, "付款节点")
    ws.cell(row, 1).font = Font(size=13, bold=True, color="1F4E78")
    row += 1
    payment_headers = ["节点", "比例", "金额", "触发条件"]
    for col, header in enumerate(payment_headers, 1):
        c = ws.cell(row, col, header)
        c.fill = header_fill
        c.font = Font(bold=True, color="1F4E78")
        c.border = border
    payments = [
        ("启动款", 0.30, f"=F{total_row}*B{row+1}", "需求范围确认、项目启动"),
        ("文字链路闭环", 0.40, f"=F{total_row}*B{row+2}", "线索、加好友、文字AI回复链路可演示"),
        ("图片链路闭环", 0.20, f"=F{total_row}*B{row+3}", "图片采集、视觉理解、图文回复链路可演示"),
        ("试运行验收", 0.10, f"=F{total_row}*B{row+4}", "灰度试运行及验收问题修复完成"),
    ]
    row += 1
    for item in payments:
        for col, value in enumerate(item, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).border = border
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    row += 2
    ws.cell(row, 1, "报价边界")
    ws.cell(row, 1).font = Font(size=13, bold=True, color="1F4E78")
    row += 1
    boundaries = [
        ("包含", "手机号线索导入、销售分配、个人微信手机号加好友、备注绑定、文字AI回复、图片另存、千问视觉理解、图文自动回复、本地车源索引、大风车基础适配、销售手动回复后AI停止、基础任务看板、日志与失败重试。"),
        ("不包含", "完整SaaS、复杂权限、计费、多商户隔离、抖音官方API、复杂BI报表、模型微调、高可用部署、长期运维、大风车商务开通费用、模型调用费、短信费、云资源费。"),
        ("前提", "甲方及时提供微信测试电脑、测试销售微信、小风车线索样本、大风车参数、千问视觉API Key、销售接管规则、图片测试样本。"),
    ]
    for label, text in boundaries:
        ws.cell(row, 1, label)
        ws.cell(row, 2, text)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        for col in range(1, 7):
            ws.cell(row, col).border = border
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 1).font = Font(bold=True)
        row += 1

    for col, width in {"A": 22, "B": 28, "C": 48, "D": 12, "E": 10, "F": 14}.items():
        ws.column_dimensions[col].width = width
    for r in range(9, total_row + 1):
        ws.row_dimensions[r].height = 34
    for r in range(total_row + 5, row + 1):
        ws.row_dimensions[r].height = 36

    for r in range(9, total_row + 1):
        ws.cell(r, 4).number_format = '"¥"#,##0'
        ws.cell(r, 5).number_format = "0.00"
        ws.cell(r, 6).number_format = '"¥"#,##0'

    for r in range(total_row + 5, total_row + 9):
        ws.cell(r, 2).number_format = "0%"
        ws.cell(r, 3).number_format = '"¥"#,##0'

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=False)
    dv.error = "人天必须为数字"
    dv.errorTitle = "输入错误"
    ws.add_data_validation(dv)
    dv.add(f"E{start_row}:E{total_row-1}")

    summary = wb.create_sheet("总览")
    summary["A1"] = "AI智能客服售前跟进系统 报价总览"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = title_fill
    summary.merge_cells("A1:E1")
    summary["A3"] = "总工程量"
    summary["B3"] = f"='报价明细'!E{total_row}"
    summary["A4"] = "人天单价"
    summary["B4"] = UNIT_PRICE
    summary["A5"] = "项目总价"
    summary["B5"] = f"='报价明细'!F{total_row}"
    summary["A6"] = "周期"
    summary["B6"] = "6-8周"
    for cell in ("A3", "A4", "A5", "A6"):
        summary[cell].font = Font(bold=True)
    summary["B4"].number_format = '"¥"#,##0'
    summary["B5"].number_format = '"¥"#,##0'
    summary["B3"].number_format = "0.00"
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 24

    summary["A8"] = "模块"
    summary["B8"] = "人天"
    summary["C8"] = "金额"
    for c in summary[8]:
        c.fill = header_fill
        c.font = Font(bold=True, color="1F4E78")
        c.border = border
    module_totals = {}
    for module, _, days, _ in QUOTE_ROWS:
        module_totals[module] = module_totals.get(module, 0) + days
    r = 9
    for module, days in module_totals.items():
        summary.cell(r, 1, module)
        summary.cell(r, 2, days)
        summary.cell(r, 3, f"=B{r}*B4")
        for col in range(1, 4):
            summary.cell(r, col).border = border
            summary.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)
        summary.cell(r, 2).number_format = "0.00"
        summary.cell(r, 3).number_format = '"¥"#,##0'
        r += 1

    wb.save(XLSX_PATH)

    # Load once to verify it is a valid workbook.
    loaded = load_workbook(XLSX_PATH, data_only=False)
    assert "报价明细" in loaded.sheetnames
    assert "总览" in loaded.sheetnames


class ArchitectureDiagram(Flowable):
    def __init__(self, width=165 * mm, height=88 * mm):
        super().__init__()
        self.width = width
        self.height = height

    def draw(self):
        c = self.canv
        c.setStrokeColor(colors.HexColor("#2E74B5"))
        c.setFillColor(colors.HexColor("#EAF2F8"))
        c.setLineWidth(1)
        boxes = [
            (0, 58, 38, 16, "抖音线索"),
            (45, 58, 42, 16, "业务控制面"),
            (94, 58, 42, 16, "RPA Worker"),
            (142, 58, 24, 16, "微信"),
            (45, 28, 42, 16, "OmniAuto"),
            (94, 28, 42, 16, "图片/视觉"),
            (142, 28, 24, 16, "车源"),
            (68, 0, 48, 16, "Guard/接管"),
        ]
        for x, y, w, h, label in boxes:
            c.roundRect(x * mm, y * mm, w * mm, h * mm, 3 * mm, fill=1, stroke=1)
            c.setFillColor(colors.HexColor("#0B2545"))
            c.setFont("STSong-Light", 9)
            c.drawCentredString((x + w / 2) * mm, (y + 5.5) * mm, label)
            c.setFillColor(colors.HexColor("#EAF2F8"))
        c.setStrokeColor(colors.HexColor("#6B7A90"))
        for x1, y1, x2, y2 in [
            (38, 66, 45, 66),
            (87, 66, 94, 66),
            (136, 66, 142, 66),
            (115, 58, 115, 44),
            (87, 36, 94, 36),
            (136, 36, 142, 36),
            (66, 58, 66, 44),
            (115, 28, 105, 16),
            (66, 28, 88, 16),
        ]:
            c.line(x1 * mm, y1 * mm, x2 * mm, y2 * mm)


def p(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(text.replace("\n", "<br/>"), style)


def bullet_list(items: list[str], style: ParagraphStyle) -> ListFlowable:
    return ListFlowable(
        [ListItem(p(item, style), leftIndent=10) for item in items],
        bulletType="bullet",
        start="circle",
        leftIndent=18,
    )


def make_table(data, col_widths, repeat=0, font_size=8.5):
    table = Table(data, colWidths=col_widths, repeatRows=repeat, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2F3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def generate_pdf() -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    doc = SimpleDocTemplate(
        str(PDF_PATH),
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=f"{PROJECT_NAME} 技术方案手册",
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "CNBody",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=9.5,
        leading=14,
        spaceAfter=5,
        textColor=colors.HexColor("#1F2937"),
    )
    small = ParagraphStyle(
        "CNSmall",
        parent=body,
        fontSize=8.2,
        leading=12,
        spaceAfter=3,
    )
    title = ParagraphStyle(
        "CNTitle",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=22,
        leading=28,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0B2545"),
        spaceAfter=12,
    )
    subtitle = ParagraphStyle(
        "CNSubtitle",
        parent=body,
        fontSize=11,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=14,
    )
    h1 = ParagraphStyle(
        "CNH1",
        parent=styles["Heading1"],
        fontName="STSong-Light",
        fontSize=15,
        leading=20,
        textColor=colors.HexColor("#1F4E78"),
        spaceBefore=8,
        spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "CNH2",
        parent=styles["Heading2"],
        fontName="STSong-Light",
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#2E74B5"),
        spaceBefore=6,
        spaceAfter=4,
    )
    code = ParagraphStyle(
        "CNCode",
        parent=body,
        fontName="STSong-Light",
        fontSize=8.3,
        leading=11,
        backColor=colors.HexColor("#F4F6F9"),
        borderColor=colors.HexColor("#E5E7EB"),
        borderWidth=0.4,
        borderPadding=5,
        spaceBefore=4,
        spaceAfter=6,
    )

    story = []
    story.append(Spacer(1, 16 * mm))
    story.append(p("AI智能客服售前跟进系统", title))
    story.append(p("技术方案手册（工程交付版）", subtitle))
    cover_data = [
        ["项目定位", "抖音小风车手机号线索驱动的个人微信AI销售预跟进系统"],
        ["版本", VERSION],
        ["日期", DOC_DATE],
        ["预算", "87,000元（58人天 × 1,500元/人天）"],
        ["周期", "6-8周"],
        ["交付标准", "范围明确、过程可追踪、验收可判定、变更可管理"],
    ]
    story.append(make_table([[p(a, body), p(b, body)] for a, b in cover_data], [35 * mm, 120 * mm], font_size=9.2))
    story.append(Spacer(1, 8 * mm))
    story.append(p("本文档用于第一期正式工程版本立项、实施、验收和交付管理。若后续产品化、SaaS化或扩大销售设备规模，应另行评估架构、预算和周期。", body))
    story.append(PageBreak())

    story.append(p("1. 项目目标与交付原则", h1))
    story.append(p("本项目第一期目标是将抖音直播小风车手机号线索，通过销售个人微信完成自动加好友、AI暖场、图文自动回复、意向识别和人工接管。系统不以AI独立成交为目标，而以提升销售线索承接效率和高意向客户识别效率为目标。", body))
    story.append(p("工程交付原则", h2))
    story.append(
        bullet_list(
            [
                "范围明确：每个模块必须对应交付物、除外项和验收口径。",
                "前提明确：账号、测试环境、API Key、白名单、样本数据由甲方按时提供。",
                "过程可追踪：需求、变更、缺陷、测试证据均形成记录。",
                "验收可判定：以闭环场景和验收用例为准，不以主观“像不像真人”为唯一标准。",
                "变更可管理：超出一期正式工程范围的需求进入变更流程，评估人天、风险和排期。",
            ],
            body,
        )
    )

    story.append(p("2. 系统总体架构", h1))
    story.append(ArchitectureDiagram())
    story.append(Spacer(1, 4 * mm))
    story.append(p("系统采用“业务控制面 + 本地RPA Worker + OmniAuto AI引擎 + 图片理解 + 车源索引”的架构。业务状态由控制面保存，微信操作由Worker执行，回复生成统一进入OmniAuto，图片理解和车源检索作为证据输入，不绕过统一回复引擎。", body))
    story.append(p("主流程", code))
    story.append(p("小风车手机号线索 -> 销售分配 -> 加好友任务 -> RPA控制个人微信添加好友 -> 备注绑定 -> OmniAuto监听文字/图片 -> AI自动回复 -> 高意向/风险接管 -> 销售手动回复后AI停止", code))

    story.append(p("3. 模块设计", h1))
    module_table = [
        ["模块", "职责", "验收要点"],
        ["线索与销售控制面", "管理手机号线索、销售、分配、任务状态", "可导入线索、分配销售、查看任务状态"],
        ["RPA Worker", "控制销售电脑微信，加好友、备注、监听、发送", "能按手机号搜索、提交申请、上报结果"],
        ["OmniAuto AI引擎", "统一生成文字和图文回复", "能结合上下文、知识、风格和规则回复"],
        ["图片理解模块", "保存微信图片并调用千问视觉识别", "能识别图片消息并输出结构化结果"],
        ["车源服务", "同步大风车车源并建立本地索引", "能按本地索引返回可见车源字段"],
        ["Guard风控", "限制价格、车况、金融、合同等越界回复", "高风险问题进入人工接管"],
        ["人工接管", "销售手动回复后停止AI", "销售发言后该会话AI不再自动回复"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in module_table], [35 * mm, 75 * mm, 55 * mm], repeat=1))

    story.append(p("4. 关键业务规则", h1))
    story.append(p("微信备注规则", h2))
    story.append(p("备注名采用短码绑定，避免销售后续修改备注导致绑定完全失效。建议格式如下：", body))
    story.append(p("CJ-张三-A7K9-1234；转人工后可变更为：【转人工】CJ-张三-A7K9-1234。系统优先识别线索短码与手机号后四位。", code))
    story.append(p("AI停止规则", h2))
    story.append(bullet_list(["客户明确拒绝继续沟通时停止。", "达到配置的最大回复轮次时停止。", "命中高风险接管规则时停止或进入人工接管。", "检测到销售本人在该会话手动发送消息后停止。"], body))

    story.append(p("5. 图文回复链路", h1))
    story.append(p("图片消息不能绕过OmniAuto直接由视觉模型回复。视觉模型只负责“看见了什么”，OmniAuto负责“销售该怎么说”。", body))
    story.append(p("客户图片 -> RPA点开并另存 -> 图片资产入库 -> 千问视觉识别 -> ImageIntent结构化 -> 车源索引检索 -> multimodal evidence pack -> OmniAuto生成回复 -> Guard质检 -> 微信发送", code))
    evidence_table = [
        ["字段", "说明"],
        ["message_type", "image"],
        ["image_summary", "视觉模型对图片的摘要"],
        ["detected_entities", "车型、颜色、车身类型、截图内容等结构化要素"],
        ["customer_intent", "找类似车、问价格、问配置、发截图咨询等"],
        ["matched_cars", "本地车源索引返回的候选车辆"],
        ["risk_flags", "价格、事故、金融、合同等风险标记"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in evidence_table], [45 * mm, 110 * mm], repeat=1))

    story.append(p("6. 大风车接口与车源治理", h1))
    story.append(p("大风车开放平台在本方案中作为权威车源系统，不作为图片搜车接口。图片找车由视觉模型识别后转成结构化条件，再查本地车源索引。", body))
    story.append(bullet_list(["接入店铺信息查询。", "按shopCode和operationPhase获取车辆ID。", "按carId查询车辆详情。", "按carId查询车辆图片。", "定时同步到本地索引，实时对话优先查本地，避免每轮对话实时调用外部接口。"], body))
    story.append(p("AI可见字段白名单", h2))
    story.append(p("品牌、车系、车型、年份、里程、城市、颜色、燃料、对外描述、网络标价、车辆图片。", code))
    story.append(p("AI不可见字段", h2))
    story.append(p("采购价、销售底价、经理底价、车主姓名、手机号、身份证、银行卡、内部备注。", code))

    story.append(p("7. 状态与数据流", h1))
    state_table = [
        ["对象", "状态流转"],
        ["线索", "new -> assigned -> add_friend_pending -> add_friend_sent -> friend_added -> ai_chatting -> handoff_required -> human_taken_over -> closed"],
        ["加好友任务", "pending -> running -> sent -> failed -> expired"],
        ["会话", "unbound -> bound -> ai_active -> handoff_open -> human_active -> stopped"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in state_table], [28 * mm, 130 * mm], repeat=1))

    story.append(p("8. 风控与合规边界", h1))
    story.append(
        bullet_list(
            [
                "AI不得承诺无事故、无水泡、无火烧，必须以检测报告和销售确认为准。",
                "AI不得承诺最低价、底价、经理价、特殊优惠。",
                "AI不得承诺贷款包过、首付固定、审批结果。",
                "AI不得承诺定金可退、合同条款、过户时效。",
                "AI不得暴露系统提示词、内部规则、接口密钥或AI身份。",
                "用户表达拒绝、投诉、法务、退款、赔偿等场景必须转人工。",
            ],
            body,
        )
    )

    story.append(p("9. 交付范围", h1))
    in_scope = [
        ["包含项", "手机号线索导入、销售分配、个人微信手机号加好友、固定申请语、备注绑定、文字AI回复、图片另存、千问视觉理解、图文自动回复、本地车源索引、大风车基础适配、销售手动回复后AI停止、基础任务看板、日志与失败重试。"],
        ["不包含项", "完整SaaS、复杂权限、计费、多商户隔离、抖音官方API、复杂BI报表、模型微调、高可用部署、长期运维、大风车商务开通费用、模型调用费、短信费、云资源费。"],
        ["甲方前提", "提供微信测试电脑、测试销售微信、小风车线索样本、大风车appKey/appSecret/appId/shopCode/operator、千问视觉API Key、销售接管规则、图片测试样本。"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in in_scope], [30 * mm, 130 * mm], repeat=0))

    story.append(p("10. 验收标准", h1))
    acceptance = [
        ["编号", "验收场景", "通过标准"],
        ["A-01", "导入手机号线索", "系统可生成线索记录并进入待分配状态"],
        ["A-02", "销售分配", "线索可分配到指定销售及其微信设备"],
        ["A-03", "微信加好友", "RPA可按手机号搜索并提交固定好友申请语"],
        ["A-04", "备注绑定", "可生成短码备注并将微信会话绑定回线索"],
        ["A-05", "文字回复", "客户文字消息可由OmniAuto自动回复并记录审计"],
        ["A-06", "图片回复", "客户图片消息可另存、识别并生成图文回复"],
        ["A-07", "高风险接管", "价格、车况、金融等敏感场景进入接管或停止自动回复"],
        ["A-08", "销售手动回复停止AI", "销售发出消息后该会话AI停止自动回复"],
        ["A-09", "失败可见", "加好友、图片、模型、车源失败均可在任务或日志中查看"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in acceptance], [16 * mm, 54 * mm, 88 * mm], repeat=1, font_size=8.0))

    story.append(p("11. 缺陷等级与检收规则", h1))
    defect = [
        ["等级", "定义", "处理要求"],
        ["S1", "核心链路不可用，如无法加好友、无法监听、无法发送回复", "阻塞验收，优先修复"],
        ["S2", "核心功能部分失败，如图片链路偶发失败、接管状态错误", "验收前修复或双方确认规避方案"],
        ["S3", "非核心缺陷，如界面文案、轻微日志展示问题", "不阻塞验收，进入遗留问题清单"],
        ["S4", "优化建议或后续增强需求", "不计入本期缺陷，进入变更或二期需求"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in defect], [18 * mm, 82 * mm, 58 * mm], repeat=1))

    story.append(p("12. 变更控制", h1))
    story.append(
        bullet_list(
            [
                "本手册、报价单和双方确认的验收清单构成本期正式工程范围基线。",
                "新增抖音官方API、多商户隔离、模型微调、复杂BI、长期运维等均视为范围变更。",
                "范围变更需记录变更内容、原因、影响模块、追加人天、交付日期影响及确认人。",
                "未经确认的口头需求不作为验收依据。",
            ],
            body,
        )
    )

    story.append(p("13. 计划与里程碑", h1))
    milestones = [
        ["阶段", "周期", "交付物"],
        ["P1 启动与基础闭环", "第1-2周", "线索导入、销售分配、加好友任务、备注规则"],
        ["P2 文字AI闭环", "第3-4周", "文字AI回复、上下文绑定、停止规则、审计"],
        ["P3 图片与图文回复", "第5-6周", "图片另存、视觉识别、图文回复、回归样本"],
        ["P4 车源与灰度验收", "第7-8周", "大风车基础适配、本地索引、灰度测试、交付文档"],
    ]
    story.append(make_table([[p(str(c), small) for c in row] for row in milestones], [32 * mm, 28 * mm, 98 * mm], repeat=1))

    story.append(p("14. 待办事项", h1))
    story.append(
        bullet_list(
            [
                "确认大风车operationPhase枚举和值含义。",
                "确认appKey、appSecret、appId、shopCode、operator。",
                "确认服务器出口IP白名单。",
                "确认销售接管规则和销售手动回复停止AI的最终口径。",
                "收集真实图片咨询样本，用于图文回复训练与回归测试。",
            ],
            body,
        )
    )

    story.append(p("15. 技术结论", h1))
    story.append(p("本方案不把系统做成单纯微信RPA脚本，而是将业务状态、微信操作、AI回复、图片理解、车源事实、风控接管拆成清晰边界。第一期可满足正式工程闭环，后续若产品化，应在现有边界上另行设计多租户、权限、计费、运维和渠道能力。", body))

    def header_footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("STSong-Light", 8)
        canvas.setFillColor(colors.HexColor("#6B7280"))
        canvas.drawString(18 * mm, 287 * mm, f"{PROJECT_NAME} 技术方案手册 {VERSION}")
        canvas.drawRightString(192 * mm, 10 * mm, f"第 {canvas.getPageNumber()} 页")
        canvas.restoreState()

    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    generate_xlsx()
    generate_pdf()
    print(PDF_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
