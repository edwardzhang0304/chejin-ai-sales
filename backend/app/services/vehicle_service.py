from __future__ import annotations

from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from pathlib import Path
import re
import warnings
from uuid import uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, UnidentifiedImageError
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.request_context import ActorContext
from app.errors import AppError
from app.models.base import utcnow
from app.models.vehicle import (
    KnowledgeCategory,
    KnowledgeItem,
    KnowledgeTenant,
    VehicleFileCleanup,
    VehicleImage,
    VehicleImportPreview,
)
from app.schemas.vehicle import VehicleCreate, VehicleUpdate
from app.services.audit_service import record_vehicle_operation_failure, write_log


PRODUCT_LAYER = "product_master"
PRODUCT_CATEGORY = "products"
EXCEL_TEMPLATE_VERSION = "1.1"
VEHICLE_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,127}$")
IMAGE_TYPES = {
    "jpeg": ("image/jpeg", ".jpg"),
    "png": ("image/png", ".png"),
    "webp": ("image/webp", ".webp"),
}
EXCEL_HEADERS = [
    "车辆编号",
    "展示名称",
    "品牌",
    "车系",
    "车型",
    "公开售价",
    "首次上牌",
    "里程公里",
    "外观颜色",
    "内饰颜色",
    "所在地",
    "客户可见描述",
    "VIN",
    "车牌号",
    "采购价",
    "内部备注",
]
EXCEL_FIELD_MAP = dict(
    zip(
        EXCEL_HEADERS,
        [
            "vehicle_code",
            "display_name",
            "brand",
            "series",
            "model",
            "public_price",
            "first_registration",
            "mileage_km",
            "exterior_color",
            "interior_color",
            "location",
            "customer_description",
            "vin",
            "plate_number",
            "purchase_price",
            "internal_notes",
        ],
    )
)


def _tenant_id() -> str:
    return get_settings().omniauto_knowledge_tenant


def _invalidate_pending_vehicle_replies(db: Session, vehicle_id: str) -> list[str]:
    # Local import keeps Product Master independent from C3 module import order.
    from app.services.c3_service import invalidate_vehicle_dependent_reply_actions

    return invalidate_vehicle_dependent_reply_actions(db, vehicle_id)


def knowledge_runtime_readiness(db: Session) -> dict:
    item_count = db.scalar(
        select(func.count()).select_from(KnowledgeItem).where(KnowledgeItem.tenant_id == _tenant_id())
    ) or 0
    return {
        "ready": True,
        "backend": "postgres" if not get_settings().database_url.startswith("sqlite") else "sqlite_test",
        "tenant_id": _tenant_id(),
        "schema": get_settings().omniauto_knowledge_schema,
        "knowledge_item_count": item_count,
    }


def _vehicle_query():
    return select(KnowledgeItem).where(
        KnowledgeItem.tenant_id == _tenant_id(),
        KnowledgeItem.layer == PRODUCT_LAYER,
        KnowledgeItem.category_id == PRODUCT_CATEGORY,
        KnowledgeItem.product_id == "",
    )


def _vehicle_or_404(db: Session, vehicle_id: str, *, lock: bool = False) -> KnowledgeItem:
    if not VEHICLE_ID_RE.fullmatch(vehicle_id):
        raise AppError("VEHICLE_ID_INVALID", "车辆编号格式不正确", 400)
    query = _vehicle_query().where(KnowledgeItem.item_id == vehicle_id)
    if lock:
        query = query.with_for_update()
    row = db.scalar(query)
    if not row:
        raise AppError("VEHICLE_NOT_FOUND", "车辆不存在", 404)
    return row


def _images(db: Session, vehicle_id: str) -> list[VehicleImage]:
    return list(
        db.scalars(
            select(VehicleImage)
            .where(VehicleImage.tenant_id == _tenant_id(), VehicleImage.vehicle_id == vehicle_id)
            .order_by(VehicleImage.sort_order, VehicleImage.created_at, VehicleImage.id)
        )
    )


def _image_dict(item: VehicleImage) -> dict:
    return {
        "id": item.id,
        "url": f"/api/vehicles/images/{item.id}",
        "original_filename": item.original_filename,
        "content_type": item.content_type,
        "size_bytes": item.size_bytes,
        "sha256": item.sha256,
        "sort_order": item.sort_order,
        "is_main": item.sort_order == 0,
        "created_at": item.created_at,
    }


def _payload_fields(payload: dict) -> dict:
    data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    details = data.get("additional_details") if isinstance(data.get("additional_details"), dict) else {}
    internal = payload.get("internal") if isinstance(payload.get("internal"), dict) else {}
    return {
        "display_name": str(data.get("name") or ""),
        "brand": details.get("brand"),
        "series": details.get("series"),
        "model": details.get("model"),
        "public_price": data.get("price"),
        "first_registration": details.get("first_registration"),
        "mileage_km": details.get("mileage_km"),
        "exterior_color": details.get("exterior_color"),
        "interior_color": details.get("interior_color"),
        "location": details.get("location"),
        "customer_description": details.get("customer_description"),
        "vin": internal.get("vin"),
        "plate_number": internal.get("plate_number"),
        "purchase_price": internal.get("purchase_price"),
        "internal_notes": internal.get("internal_notes"),
    }


def _vehicle_dict(db: Session, row: KnowledgeItem, *, include_internal: bool = True) -> dict:
    fields = _payload_fields(row.payload or {})
    if not include_internal:
        for key in ("vin", "plate_number", "purchase_price", "internal_notes"):
            fields.pop(key, None)
    images = [_image_dict(item) for item in _images(db, row.item_id)]
    return {
        "vehicle_code": row.item_id,
        **fields,
        "listing_status": "listed" if row.status != "archived" else "unlisted",
        "images": images,
        "main_image": images[0] if images else None,
        "created_at": row.created_at,
        "updated_at": row.updated_at,
    }


def _number(value):
    if value is None:
        return None
    return float(value) if isinstance(value, Decimal) else value


def _build_product_payload(
    vehicle_id: str,
    fields: dict,
    *,
    listed: bool,
    previous: dict | None = None,
    actor: ActorContext | None = None,
) -> dict:
    now = utcnow().isoformat(timespec="seconds")
    aliases = [fields.get("brand"), fields.get("series"), fields.get("model")]
    aliases = list(dict.fromkeys(str(item).strip() for item in aliases if str(item or "").strip()))
    details = {
        key: fields.get(key)
        for key in (
            "brand",
            "series",
            "model",
            "first_registration",
            "mileage_km",
            "exterior_color",
            "interior_color",
            "location",
            "customer_description",
        )
        if fields.get(key) not in (None, "")
    }
    internal = {
        key: _number(fields.get(key))
        for key in ("vin", "plate_number", "purchase_price", "internal_notes")
        if fields.get(key) not in (None, "")
    }
    old_metadata = previous.get("metadata") if isinstance((previous or {}).get("metadata"), dict) else {}
    operator_id = str(actor.operator_id) if actor else "system"
    return {
        "schema_version": 1,
        "category_id": PRODUCT_CATEGORY,
        "id": vehicle_id,
        "status": "active" if listed else "archived",
        "source": {"type": "chejin_backend"},
        "data": {
            "name": fields["display_name"],
            "sku": vehicle_id,
            "category": "二手车",
            "aliases": aliases,
            "specs": " / ".join(aliases),
            "price": _number(fields.get("public_price")),
            "unit": "台",
            "inventory": 1,
            "additional_details": details,
        },
        "runtime": {
            "allow_auto_reply": listed,
            "requires_handoff": False,
            "risk_level": "normal",
        },
        "review_state": {
            "is_new": False,
            "acknowledged_at": now,
            "acknowledged_by": operator_id,
        },
        "internal": internal,
        "metadata": {
            "created_at": old_metadata.get("created_at") or now,
            "updated_at": now,
            "created_by": old_metadata.get("created_by") or operator_id,
            "updated_by": operator_id,
        },
    }


def _search_text(payload: dict) -> str:
    fields = _payload_fields(payload)
    return " ".join(
        str(value)
        for value in (
            payload.get("id"),
            fields.get("display_name"),
            fields.get("brand"),
            fields.get("series"),
            fields.get("model"),
            fields.get("customer_description"),
        )
        if value not in (None, "")
    )


def _ensure_catalog(db: Session) -> None:
    tenant_id = _tenant_id()
    if not db.get(KnowledgeTenant, tenant_id):
        db.add(KnowledgeTenant(tenant_id=tenant_id, display_name="车金", payload={"source": "chejin_backend"}))
    key = {"tenant_id": tenant_id, "layer": PRODUCT_LAYER, "category_id": PRODUCT_CATEGORY}
    if not db.get(KnowledgeCategory, key):
        db.add(
            KnowledgeCategory(
                **key,
                enabled=True,
                sort_order=10,
                payload={
                    "id": PRODUCT_CATEGORY,
                    "name": "车辆主数据",
                    "kind": "product_master",
                    "enabled": True,
                    "participates_in_reply": True,
                    "scope": PRODUCT_LAYER,
                    "authority": "manual_product_master_only",
                },
            )
        )


def _new_vehicle_id(db: Session) -> str:
    for _ in range(5):
        candidate = f"CJ{datetime.now().strftime('%Y%m%d')}{uuid4().hex[:8].upper()}"
        if not db.scalar(_vehicle_query().where(KnowledgeItem.item_id == candidate)):
            return candidate
    raise AppError("VEHICLE_CODE_GENERATION_FAILED", "车辆编号生成失败，请重试", 503)


def list_vehicles(
    db: Session,
    *,
    keyword: str | None,
    listing_status: str | None,
    page: int,
    page_size: int,
) -> dict:
    query = _vehicle_query()
    if listing_status == "listed":
        query = query.where(KnowledgeItem.status != "archived")
    elif listing_status == "unlisted":
        query = query.where(KnowledgeItem.status == "archived")
    elif listing_status not in (None, "", "all"):
        raise AppError("VEHICLE_LISTING_STATUS_INVALID", "上架状态不正确", 400)
    if keyword and keyword.strip():
        token = f"%{keyword.strip()}%"
        query = query.where(or_(KnowledgeItem.item_id.ilike(token), KnowledgeItem.search_text.ilike(token)))
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    rows = list(
        db.scalars(
            query.order_by(KnowledgeItem.updated_at.desc(), KnowledgeItem.item_id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return {
        "items": [_vehicle_dict(db, row) for row in rows],
        "page": page,
        "page_size": page_size,
        "total": total,
    }


def get_vehicle(db: Session, vehicle_id: str) -> dict:
    return _vehicle_dict(db, _vehicle_or_404(db, vehicle_id))


def _create_with_id(db: Session, vehicle_id: str, fields: dict, actor: ActorContext) -> KnowledgeItem:
    _ensure_catalog(db)
    payload = _build_product_payload(vehicle_id, fields, listed=False, actor=actor)
    row = KnowledgeItem(
        tenant_id=_tenant_id(),
        layer=PRODUCT_LAYER,
        category_id=PRODUCT_CATEGORY,
        product_id="",
        item_id=vehicle_id,
        status="archived",
        search_text=_search_text(payload),
        payload=payload,
    )
    db.add(row)
    db.flush()
    return row


def create_vehicle(db: Session, payload: VehicleCreate, actor: ActorContext) -> dict:
    vehicle_id = _new_vehicle_id(db)
    fields = payload.model_dump(mode="python")
    row = _create_with_id(db, vehicle_id, fields, actor)
    write_log(
        db,
        actor,
        event_type="vehicle_created",
        module="vehicles",
        target_type="vehicle",
        target_id=vehicle_id,
        after_data={"vehicle_code": vehicle_id, "listing_status": "unlisted"},
        metadata={"changed_fields": sorted(payload.model_fields_set)},
    )
    return _vehicle_dict(db, row)


def update_vehicle(db: Session, vehicle_id: str, payload: VehicleUpdate, actor: ActorContext) -> dict:
    row = _vehicle_or_404(db, vehicle_id, lock=True)
    old_payload = dict(row.payload or {})
    fields = _payload_fields(old_payload)
    changes = payload.model_dump(mode="python", exclude_unset=True)
    fields.update(changes)
    listed = row.status != "archived"
    if listed and fields.get("public_price") in (None, "", 0, Decimal("0")):
        raise AppError("VEHICLE_LISTED_PRICE_REQUIRED", "已上架车辆不能清空公开售价", 409)
    row.payload = _build_product_payload(vehicle_id, fields, listed=listed, previous=old_payload, actor=actor)
    row.search_text = _search_text(row.payload)
    row.updated_at = utcnow()
    write_log(
        db,
        actor,
        event_type="vehicle_updated",
        module="vehicles",
        target_type="vehicle",
        target_id=vehicle_id,
        before_data={"vehicle_code": vehicle_id, "listing_status": "listed" if listed else "unlisted"},
        after_data={"vehicle_code": vehicle_id, "listing_status": "listed" if listed else "unlisted"},
        metadata={"changed_fields": sorted(changes)},
    )
    _invalidate_pending_vehicle_replies(db, vehicle_id)
    db.flush()
    return _vehicle_dict(db, row)


def set_listing(db: Session, vehicle_id: str, *, listed: bool, actor: ActorContext) -> dict:
    row = _vehicle_or_404(db, vehicle_id, lock=True)
    was_listed = row.status != "archived"
    if was_listed == listed:
        return _vehicle_dict(db, row)
    fields = _payload_fields(row.payload or {})
    if listed:
        if not fields.get("display_name"):
            raise AppError("VEHICLE_LISTING_NAME_REQUIRED", "上架前必须填写展示名称", 409)
        try:
            price = Decimal(str(fields.get("public_price") or "0"))
        except InvalidOperation:
            price = Decimal("0")
        if price <= 0:
            raise AppError("VEHICLE_LISTING_PRICE_REQUIRED", "上架前必须填写大于 0 的公开售价", 409)
        if not _images(db, vehicle_id):
            raise AppError("VEHICLE_LISTING_IMAGE_REQUIRED", "上架前必须至少上传一张车辆图片", 409)
    row.status = "active" if listed else "archived"
    row.payload = _build_product_payload(
        vehicle_id,
        fields,
        listed=listed,
        previous=row.payload or {},
        actor=actor,
    )
    row.search_text = _search_text(row.payload)
    row.updated_at = utcnow()
    write_log(
        db,
        actor,
        event_type="vehicle_listed" if listed else "vehicle_unlisted",
        module="vehicles",
        target_type="vehicle",
        target_id=vehicle_id,
        before_data={"listing_status": "listed" if was_listed else "unlisted"},
        after_data={"listing_status": "listed" if listed else "unlisted"},
    )
    _invalidate_pending_vehicle_replies(db, vehicle_id)
    db.flush()
    return _vehicle_dict(db, row)


def _decode_image_type(data: bytes) -> tuple[str, str]:
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(BytesIO(data)) as image:
                image_format = str(image.format or "").lower()
                width, height = image.size
                if image_format not in IMAGE_TYPES:
                    raise AppError("VEHICLE_IMAGE_TYPE_INVALID", "只支持 JPEG、PNG 和 WebP 图片", 400)
                if width < 1 or height < 1 or width * height > get_settings().vehicle_image_max_pixels:
                    raise AppError(
                        "VEHICLE_IMAGE_DIMENSIONS_EXCEEDED",
                        "图片像素尺寸超过限制",
                        413,
                        {"max_pixels": get_settings().vehicle_image_max_pixels},
                    )
                if int(getattr(image, "n_frames", 1) or 1) != 1:
                    raise AppError("VEHICLE_IMAGE_ANIMATED_NOT_ALLOWED", "不支持动图", 400)
                image.verify()
            with Image.open(BytesIO(data)) as decoded:
                decoded.load()
    except AppError:
        raise
    except (UnidentifiedImageError, OSError, SyntaxError, ValueError, Image.DecompressionBombError, Image.DecompressionBombWarning) as exc:
        raise AppError("VEHICLE_IMAGE_DECODE_FAILED", "图片内容损坏或无法解码", 400) from exc
    return IMAGE_TYPES[image_format]


def _storage_path(storage_key: str) -> Path:
    root = Path(get_settings().vehicle_image_storage_root).expanduser().resolve()
    path = (root / storage_key).resolve()
    if root not in path.parents:
        raise AppError("VEHICLE_IMAGE_STORAGE_KEY_INVALID", "车辆图片存储键不合法", 500)
    return path


def upload_vehicle_images(
    db: Session,
    vehicle_id: str,
    files: list[tuple[str, bytes]],
    actor: ActorContext,
) -> dict:
    _vehicle_or_404(db, vehicle_id, lock=True)
    if not files:
        raise AppError("VEHICLE_IMAGE_REQUIRED", "至少选择一张车辆图片", 400)
    existing = _images(db, vehicle_id)
    settings = get_settings()
    next_order = max((item.sort_order for item in existing), default=-1) + 1
    results = []
    created_count = 0
    for filename, data in files:
        if not data:
            results.append({"filename": filename, "ok": False, "error_code": "VEHICLE_IMAGE_EMPTY"})
            continue
        if len(data) > settings.vehicle_image_max_bytes:
            results.append({"filename": filename, "ok": False, "error_code": "VEHICLE_IMAGE_TOO_LARGE"})
            continue
        try:
            detected = _decode_image_type(data)
        except AppError as exc:
            results.append({"filename": filename, "ok": False, "error_code": exc.code})
            continue
        content_type, suffix = detected
        digest = sha256(data).hexdigest()
        duplicate = db.scalar(
            select(VehicleImage).where(
                VehicleImage.tenant_id == _tenant_id(),
                VehicleImage.vehicle_id == vehicle_id,
                VehicleImage.sha256 == digest,
            )
        )
        if duplicate:
            results.append({"filename": filename, "ok": True, "duplicated": True, "image": _image_dict(duplicate)})
            continue
        if len(existing) + created_count >= settings.vehicle_image_max_count:
            results.append({"filename": filename, "ok": False, "error_code": "VEHICLE_IMAGE_COUNT_EXCEEDED"})
            continue
        image_id = str(uuid4())
        storage_key = f"{_tenant_id()}/{vehicle_id}/{image_id}{suffix}"
        path = _storage_path(storage_key)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.tmp")
        temp_path.write_bytes(data)
        temp_path.replace(path)
        db.info.setdefault("vehicle_files_created", []).append(path)
        image = VehicleImage(
            id=image_id,
            tenant_id=_tenant_id(),
            vehicle_id=vehicle_id,
            storage_key=storage_key,
            original_filename=Path(filename or f"image{suffix}").name[:255],
            content_type=content_type,
            size_bytes=len(data),
            sha256=digest,
            sort_order=next_order,
            created_by=str(actor.operator_id),
        )
        try:
            with db.begin_nested():
                db.add(image)
                db.flush()
        except IntegrityError:
            path.unlink(missing_ok=True)
            db.info["vehicle_files_created"].remove(path)
            results.append({"filename": filename, "ok": False, "error_code": "VEHICLE_IMAGE_DUPLICATED"})
            continue
        next_order += 1
        created_count += 1
        write_log(
            db,
            actor,
            event_type="vehicle_image_uploaded",
            module="vehicles",
            target_type="vehicle",
            target_id=vehicle_id,
            after_data={"image_id": image.id, "content_type": content_type, "size_bytes": len(data)},
        )
        results.append({"filename": filename, "ok": True, "duplicated": False, "image": _image_dict(image)})
    if created_count:
        _invalidate_pending_vehicle_replies(db, vehicle_id)
    failure_codes = sorted({str(item.get("error_code") or "") for item in results if not item.get("ok")})
    if failure_codes:
        write_log(
            db,
            actor,
            event_type="vehicle_operation_failed",
            module="vehicles",
            target_type="vehicle",
            target_id=vehicle_id,
            metadata={
                "operation": "upload_images",
                "error_codes": failure_codes,
                "failed_count": sum(1 for item in results if not item.get("ok")),
            },
        )
    return {
        "items": results,
        "succeeded": sum(1 for item in results if item["ok"]),
        "failed": sum(1 for item in results if not item["ok"]),
    }


def get_vehicle_image(db: Session, image_id: str) -> tuple[VehicleImage, Path]:
    image = db.scalar(select(VehicleImage).where(VehicleImage.id == image_id, VehicleImage.tenant_id == _tenant_id()))
    if not image:
        raise AppError("VEHICLE_IMAGE_NOT_FOUND", "车辆图片不存在", 404)
    path = _storage_path(image.storage_key)
    if not path.is_file():
        raise AppError("VEHICLE_IMAGE_FILE_MISSING", "车辆图片文件缺失，请联系管理员并提供 trace_id", 500)
    return image, path


def reorder_vehicle_images(db: Session, vehicle_id: str, image_ids: list[str], actor: ActorContext) -> dict:
    _vehicle_or_404(db, vehicle_id, lock=True)
    images = _images(db, vehicle_id)
    if set(image_ids) != {item.id for item in images}:
        raise AppError("VEHICLE_IMAGE_ORDER_INCOMPLETE", "排序必须包含该车辆的全部图片且不能包含其他图片", 409)
    order_changed = image_ids != [item.id for item in images]
    by_id = {item.id: item for item in images}
    for index, image_id in enumerate(image_ids):
        by_id[image_id].sort_order = index
    write_log(
        db,
        actor,
        event_type="vehicle_image_reordered",
        module="vehicles",
        target_type="vehicle",
        target_id=vehicle_id,
        after_data={"image_ids": image_ids},
    )
    if order_changed:
        _invalidate_pending_vehicle_replies(db, vehicle_id)
    db.flush()
    return {"items": [_image_dict(by_id[image_id]) for image_id in image_ids]}


def prepare_delete_vehicle_image(db: Session, vehicle_id: str, image_id: str, actor: ActorContext) -> tuple[dict, str]:
    vehicle = _vehicle_or_404(db, vehicle_id, lock=True)
    images = _images(db, vehicle_id)
    image = next((item for item in images if item.id == image_id), None)
    if not image:
        raise AppError("VEHICLE_IMAGE_NOT_FOUND", "车辆图片不存在", 404)
    if vehicle.status != "archived" and len(images) <= 1:
        raise AppError("VEHICLE_LISTED_LAST_IMAGE_REQUIRED", "已上架车辆不能删除最后一张图片", 409)
    cleanup = VehicleFileCleanup(
        tenant_id=_tenant_id(),
        vehicle_id=vehicle_id,
        storage_key=image.storage_key,
        status="pending",
    )
    db.add(cleanup)
    db.delete(image)
    remaining = [item for item in images if item.id != image_id]
    for index, item in enumerate(remaining):
        item.sort_order = index
    write_log(
        db,
        actor,
        event_type="vehicle_image_deleted",
        module="vehicles",
        target_type="vehicle",
        target_id=vehicle_id,
        before_data={"image_id": image_id},
    )
    _invalidate_pending_vehicle_replies(db, vehicle_id)
    db.flush()
    return {"image_id": image_id, "deleted": True}, cleanup.id


def process_vehicle_file_cleanup(cleanup_id: str) -> bool:
    from app.core.database import SessionLocal

    with SessionLocal() as db:
        cleanup = db.get(VehicleFileCleanup, cleanup_id)
        if not cleanup or cleanup.status == "completed":
            return True
        cleanup.attempts += 1
        try:
            _storage_path(cleanup.storage_key).unlink(missing_ok=True)
            cleanup.status = "completed"
            cleanup.last_error = ""
            cleanup.completed_at = utcnow()
            db.commit()
            return True
        except OSError as exc:
            cleanup.status = "pending"
            cleanup.last_error = type(exc).__name__[:128]
            db.commit()
            return False


def retry_pending_vehicle_file_cleanups(*, limit: int = 100) -> dict[str, int]:
    from app.core.database import SessionLocal

    with SessionLocal() as db:
        cleanup_ids = list(
            db.scalars(
                select(VehicleFileCleanup.id)
                .where(VehicleFileCleanup.status == "pending")
                .order_by(VehicleFileCleanup.created_at, VehicleFileCleanup.id)
                .limit(max(1, limit))
            )
        )
    completed = sum(1 for cleanup_id in cleanup_ids if process_vehicle_file_cleanup(cleanup_id))
    return {"attempted": len(cleanup_ids), "completed": completed, "pending": len(cleanup_ids) - completed}


def build_excel_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "车辆信息"
    sheet.append(EXCEL_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P1"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 18
    notes = workbook.create_sheet("字段说明")
    notes.append(["字段", "说明"])
    descriptions = {
        "车辆编号": "新增必须留空并由系统生成；填写已存在编号时更新；填写不存在编号会报错",
        "展示名称": "必填",
        "公开售价": "数字，单位由业务统一解释；上架前必须大于 0",
        "首次上牌": "YYYY-MM",
        "里程公里": "非负整数",
    }
    for header in EXCEL_HEADERS:
        notes.append([header, descriptions.get(header, "留空表示不修改已有值")])
    metadata = workbook.create_sheet("_元数据")
    metadata.sheet_state = "hidden"
    metadata.append(["template_version", EXCEL_TEMPLATE_VERSION])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _validate_xlsx_archive(data: bytes) -> None:
    try:
        with ZipFile(BytesIO(data)) as archive:
            names = set(archive.namelist())
            if "xl/workbook.xml" not in names:
                raise AppError("VEHICLE_EXCEL_INVALID", "文件不是有效的 XLSX 工作簿", 400)
            expanded = sum(item.file_size for item in archive.infolist())
            if expanded > 64 * 1024 * 1024:
                raise AppError("VEHICLE_EXCEL_EXPANDED_TOO_LARGE", "Excel 解压后内容过大", 413)
            if any(name.lower().endswith(("vbaproject.bin", ".exe", ".dll")) for name in names):
                raise AppError("VEHICLE_EXCEL_UNSAFE_CONTENT", "Excel 含不允许的可执行内容", 400)
    except BadZipFile as exc:
        raise AppError("VEHICLE_EXCEL_INVALID", "文件不是有效的 XLSX 工作簿", 400) from exc


def _clean_excel_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _decimal_value(value, field_name: str, errors: list[str]):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        errors.append(f"{field_name}必须是数字")
        return None
    if number < 0:
        errors.append(f"{field_name}不能小于 0")
    return number


def _integer_value(value, field_name: str, errors: list[str]):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
        if number != number.to_integral_value():
            raise ValueError
        result = int(number)
    except (InvalidOperation, ValueError):
        errors.append(f"{field_name}必须是整数")
        return None
    if result < 0:
        errors.append(f"{field_name}不能小于 0")
    return result


def preview_excel_import(db: Session, *, filename: str, data: bytes, actor: ActorContext) -> dict:
    settings = get_settings()
    if not filename.lower().endswith(".xlsx"):
        raise AppError("VEHICLE_EXCEL_TYPE_INVALID", "只支持 .xlsx 文件", 400)
    if not data:
        raise AppError("VEHICLE_EXCEL_EMPTY", "Excel 文件为空", 400)
    if len(data) > settings.vehicle_excel_max_bytes:
        raise AppError("VEHICLE_EXCEL_TOO_LARGE", "Excel 文件超过大小限制", 413, {"max_bytes": settings.vehicle_excel_max_bytes})
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise AppError("VEHICLE_EXCEL_INVALID", "Excel 文件无法解析", 400) from exc
    if "车辆信息" not in workbook.sheetnames or "字段说明" not in workbook.sheetnames or "_元数据" not in workbook.sheetnames:
        raise AppError("VEHICLE_EXCEL_TEMPLATE_INVALID", "Excel 模板工作表不完整", 400)
    metadata = workbook["_元数据"]
    if metadata["A1"].value != "template_version" or str(metadata["B1"].value or "") != EXCEL_TEMPLATE_VERSION:
        raise AppError("VEHICLE_EXCEL_TEMPLATE_VERSION_MISMATCH", "Excel 模板版本不匹配", 409, {"expected": EXCEL_TEMPLATE_VERSION})
    sheet = workbook["车辆信息"]
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers != EXCEL_HEADERS:
        raise AppError("VEHICLE_EXCEL_HEADER_INVALID", "Excel 表头与模板不一致", 400, {"expected": EXCEL_HEADERS})
    raw_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    while raw_rows and all(_clean_excel_value(value) is None for value in raw_rows[-1]):
        raw_rows.pop()
    if len(raw_rows) > settings.vehicle_excel_max_rows:
        raise AppError("VEHICLE_EXCEL_ROWS_EXCEEDED", "Excel 数据行数超过限制", 413, {"max_rows": settings.vehicle_excel_max_rows})

    codes = [str(_clean_excel_value(row[0]) or "") for row in raw_rows]
    nonempty_codes = {code for code in codes if code}
    existing_rows = list(db.scalars(_vehicle_query().where(KnowledgeItem.item_id.in_(nonempty_codes)))) if nonempty_codes else []
    existing = {item.item_id: item for item in existing_rows}
    seen: set[str] = set()
    preview_rows = []
    create_count = update_count = error_count = 0
    for index, values in enumerate(raw_rows, start=2):
        errors: list[str] = []
        raw = {EXCEL_FIELD_MAP[header]: _clean_excel_value(value) for header, value in zip(EXCEL_HEADERS, values)}
        vehicle_id = str(raw.pop("vehicle_code") or "")
        action = "update" if vehicle_id else "create"
        if vehicle_id:
            if not VEHICLE_ID_RE.fullmatch(vehicle_id):
                errors.append("车辆编号格式不正确")
            elif vehicle_id not in existing:
                errors.append("车辆编号不存在；如需新增请清空车辆编号，由系统生成")
            if vehicle_id in seen:
                errors.append("同一文件内车辆编号重复")
            seen.add(vehicle_id)
        else:
            vehicle_id = _new_vehicle_id(db)
        raw["public_price"] = _decimal_value(raw.get("public_price"), "公开售价", errors)
        raw["purchase_price"] = _decimal_value(raw.get("purchase_price"), "采购价", errors)
        raw["mileage_km"] = _integer_value(raw.get("mileage_km"), "里程公里", errors)
        if raw.get("first_registration") and not re.fullmatch(r"^(19|20)\d{2}-(0[1-9]|1[0-2])$", str(raw["first_registration"])):
            errors.append("首次上牌必须为 YYYY-MM")
        if action == "create" and not raw.get("display_name"):
            errors.append("新增车辆必须填写展示名称")
        changes = {key: value for key, value in raw.items() if value is not None}
        if action == "update" and not changes:
            errors.append("已有车辆没有可更新字段")
        try:
            if action == "create":
                VehicleCreate.model_validate(changes)
            elif changes:
                VehicleUpdate.model_validate(changes)
        except Exception as exc:
            errors.append(str(exc))
        if errors:
            error_count += 1
        elif action == "create":
            create_count += 1
        else:
            update_count += 1
        preview_rows.append(
            {
                "row_number": index,
                "vehicle_code": vehicle_id,
                "action": action,
                "data": {key: _number(value) for key, value in changes.items()},
                "errors": errors,
            }
        )
    result = {
        "total_rows": len(preview_rows),
        "create_count": create_count,
        "update_count": update_count,
        "error_count": error_count,
        "can_confirm": error_count == 0 and bool(preview_rows),
        "rows": preview_rows,
    }
    preview = VehicleImportPreview(
        tenant_id=_tenant_id(),
        template_version=EXCEL_TEMPLATE_VERSION,
        file_name=Path(filename).name[:255],
        file_sha256=sha256(data).hexdigest(),
        status="pending",
        rows_payload=preview_rows,
        result_payload=result,
        created_by=str(actor.operator_id),
        expires_at=utcnow() + timedelta(seconds=settings.vehicle_import_preview_ttl_seconds),
    )
    db.add(preview)
    db.flush()
    return {"preview_id": preview.id, "status": preview.status, "expires_at": preview.expires_at, **result}


def confirm_excel_import(db: Session, preview_id: str, actor: ActorContext) -> dict:
    preview = db.scalar(
        select(VehicleImportPreview)
        .where(VehicleImportPreview.id == preview_id, VehicleImportPreview.tenant_id == _tenant_id())
        .with_for_update()
    )
    if not preview:
        raise AppError("VEHICLE_IMPORT_PREVIEW_NOT_FOUND", "导入预览不存在", 404)
    if preview.status == "confirmed":
        return {"preview_id": preview.id, "status": "confirmed", "duplicated": True, **(preview.result_payload or {})}
    now = utcnow()
    expires_at = preview.expires_at
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at <= now:
        preview.status = "expired"
        raise AppError("VEHICLE_IMPORT_PREVIEW_EXPIRED", "导入预览已过期，请重新预览", 409)
    result = dict(preview.result_payload or {})
    if int(result.get("error_count") or 0) > 0 or not result.get("can_confirm"):
        raise AppError("VEHICLE_IMPORT_HAS_ERRORS", "导入预览存在错误，不能确认", 409, {"error_count": result.get("error_count", 0)})
    for item in preview.rows_payload or []:
        vehicle_id = str(item["vehicle_code"])
        changes = dict(item.get("data") or {})
        if item.get("action") == "create":
            if db.scalar(_vehicle_query().where(KnowledgeItem.item_id == vehicle_id)):
                raise AppError("VEHICLE_IMPORT_CONCURRENT_CONFLICT", "预览后车辆编号已被占用，请重新预览", 409, {"vehicle_code": vehicle_id})
            validated = VehicleCreate.model_validate(changes)
            _create_with_id(db, vehicle_id, validated.model_dump(mode="python"), actor)
        else:
            row = _vehicle_or_404(db, vehicle_id, lock=True)
            fields = _payload_fields(row.payload or {})
            fields.update(VehicleUpdate.model_validate(changes).model_dump(mode="python", exclude_unset=True))
            listed = row.status != "archived"
            if listed and Decimal(str(fields.get("public_price") or "0")) <= 0:
                raise AppError("VEHICLE_IMPORT_LISTED_PRICE_REQUIRED", "导入不能清空已上架车辆的公开售价", 409, {"vehicle_code": vehicle_id})
            row.payload = _build_product_payload(vehicle_id, fields, listed=listed, previous=row.payload or {}, actor=actor)
            row.search_text = _search_text(row.payload)
            row.updated_at = utcnow()
            _invalidate_pending_vehicle_replies(db, vehicle_id)
    preview.status = "confirmed"
    preview.confirmed_at = now
    result.update({"confirmed_at": now.isoformat(), "imported_count": int(result.get("create_count") or 0) + int(result.get("update_count") or 0)})
    preview.result_payload = result
    write_log(
        db,
        actor,
        event_type="vehicle_excel_import_confirmed",
        module="vehicles",
        target_type="vehicle_import",
        target_id=preview.id,
        after_data={
            "file_sha256": preview.file_sha256,
            "create_count": result.get("create_count", 0),
            "update_count": result.get("update_count", 0),
        },
    )
    db.flush()
    return {"preview_id": preview.id, "status": preview.status, "duplicated": False, **result}
