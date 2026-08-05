import base64
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select
from uuid import UUID

from app.core.config import get_settings
from app.core.request_context import ActorContext
from app.core.database import Base, SessionLocal, engine
from app.main import app
from app.models.audit import OperationLog
from app.models.vehicle import KnowledgeItem, VehicleFileCleanup, VehicleImage, VehicleImportPreview
from app.services import vehicle_service


client = TestClient(app)
ADMIN_HEADERS = {
    "X-Operator-Id": "00000000-0000-0000-0000-000000000001",
    "X-Operator-Name": "Vehicle Admin",
    "X-Operator-Role": "admin",
}
SALES_HEADERS = {
    "Authorization": "Bearer sales-reader-token",
    "X-Operator-Id": "00000000-0000-0000-0000-000000000002",
    "X-Operator-Name": "Sales Reader",
    "X-Operator-Role": "admin",
}
def _image_bytes(image_format: str, color: tuple[int, int, int]) -> bytes:
    output = BytesIO()
    Image.new("RGB", (2, 2), color).save(output, format=image_format)
    return output.getvalue()


PNG_1X1 = _image_bytes("PNG", (255, 0, 0))
JPEG_MINIMAL = _image_bytes("JPEG", (0, 255, 0))
WEBP_MINIMAL = _image_bytes("WEBP", (0, 0, 255))


def setup_function():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def _create_vehicle(**overrides) -> dict:
    payload = {"display_name": "2022款测试轿车", "brand": "测试品牌", "series": "测试车系"}
    payload.update(overrides)
    response = client.post("/api/vehicles", json=payload, headers=ADMIN_HEADERS)
    assert response.status_code == 200, response.text
    return response.json()["data"]


def test_vehicle_product_master_payload_excludes_internal_fields_from_brain_projection():
    vehicle = _create_vehicle(
        vin="LTEST123456789012",
        plate_number="苏A12345",
        purchase_price=7.5,
        internal_notes="内部底价不可外发",
    )

    assert vehicle["listing_status"] == "unlisted"
    assert vehicle["vehicle_code"].startswith("CJ")
    with SessionLocal() as db:
        item = db.scalar(select(KnowledgeItem).where(KnowledgeItem.item_id == vehicle["vehicle_code"]))
        assert item is not None
        assert item.layer == "product_master"
        assert item.category_id == "products"
        assert item.status == "archived"
        assert item.payload["data"]["name"] == "2022款测试轿车"
        assert "vin" not in item.payload["data"]
        assert "purchase_price" not in item.payload["data"].get("additional_details", {})
        assert "LTEST123456789012" not in item.search_text
        assert item.payload["internal"]["vin"] == "LTEST123456789012"
        log = db.scalar(select(OperationLog).where(OperationLog.event_type == "vehicle_created"))
        assert log is not None
        assert "LTEST123456789012" not in str(log.after_data)


def test_listing_requires_price_and_image_then_brain_catalog_becomes_active():
    vehicle = _create_vehicle()
    vehicle_id = vehicle["vehicle_code"]

    missing_price = client.post(f"/api/vehicles/{vehicle_id}/list", headers=ADMIN_HEADERS)
    assert missing_price.status_code == 409
    assert missing_price.json()["code"] == "VEHICLE_LISTING_PRICE_REQUIRED"

    updated = client.put(
        f"/api/vehicles/{vehicle_id}",
        json={"public_price": 12.88, "mileage_km": 23000, "first_registration": "2022-03"},
        headers=ADMIN_HEADERS,
    )
    assert updated.status_code == 200
    missing_image = client.post(f"/api/vehicles/{vehicle_id}/list", headers=ADMIN_HEADERS)
    assert missing_image.status_code == 409
    assert missing_image.json()["code"] == "VEHICLE_LISTING_IMAGE_REQUIRED"

    uploaded = client.post(
        f"/api/vehicles/{vehicle_id}/images",
        files=[
            ("files", ("bad.txt", b"not-an-image", "text/plain")),
            ("files", ("car.png", PNG_1X1, "image/png")),
        ],
        headers=ADMIN_HEADERS,
    )
    assert uploaded.status_code == 200, uploaded.text
    upload_data = uploaded.json()["data"]
    assert upload_data["succeeded"] == 1
    assert upload_data["failed"] == 1

    listed = client.post(f"/api/vehicles/{vehicle_id}/list", headers=ADMIN_HEADERS)
    assert listed.status_code == 200
    assert listed.json()["data"]["listing_status"] == "listed"
    image_id = listed.json()["data"]["images"][0]["id"]
    last_delete = client.delete(f"/api/vehicles/{vehicle_id}/images/{image_id}", headers=ADMIN_HEADERS)
    assert last_delete.status_code == 409
    assert last_delete.json()["code"] == "VEHICLE_LISTED_LAST_IMAGE_REQUIRED"

    with SessionLocal() as db:
        item = db.scalar(select(KnowledgeItem).where(KnowledgeItem.item_id == vehicle_id))
        assert item.status == "active"
        assert item.payload["runtime"]["allow_auto_reply"] is True

    assert client.post(f"/api/vehicles/{vehicle_id}/unlist", headers=ADMIN_HEADERS).status_code == 200
    assert client.delete(f"/api/vehicles/{vehicle_id}/images/{image_id}", headers=ADMIN_HEADERS).status_code == 200


def test_every_authenticated_account_can_read_and_modify_vehicle_catalog():
    vehicle = _create_vehicle()
    listed = client.get("/api/vehicles", headers=SALES_HEADERS)
    assert listed.status_code == 200
    updated = client.put(
        f"/api/vehicles/{vehicle['vehicle_code']}",
        json={"brand": "同权限账号可修改"},
        headers=SALES_HEADERS,
    )
    assert updated.status_code == 200
    assert updated.json()["data"]["brand"] == "同权限账号可修改"


def test_excel_preview_confirm_is_atomic_and_idempotent():
    template = client.get("/api/vehicles/excel/template", headers=ADMIN_HEADERS)
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["车辆信息"]
    sheet.append([None, "Excel新增车辆", "品牌A", "车系A", "车型A", 9.98, "2021-08", 31000])
    content = BytesIO()
    workbook.save(content)

    preview = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("vehicles.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    )
    assert preview.status_code == 200, preview.text
    preview_data = preview.json()["data"]
    assert preview_data["can_confirm"] is True
    assert preview_data["create_count"] == 1

    endpoint = f"/api/vehicles/excel/{preview_data['preview_id']}/confirm"
    confirmed = client.post(endpoint, headers=ADMIN_HEADERS)
    repeated = client.post(endpoint, headers=ADMIN_HEADERS)
    assert confirmed.status_code == 200
    assert confirmed.json()["data"]["imported_count"] == 1
    assert repeated.status_code == 200
    assert repeated.json()["data"]["duplicated"] is True

    vehicles = client.get("/api/vehicles?keyword=Excel新增车辆", headers=ADMIN_HEADERS).json()["data"]
    assert vehicles["total"] == 1
    assert vehicles["items"][0]["listing_status"] == "unlisted"


def test_excel_errors_block_confirm_without_writing_vehicle_facts():
    template = client.get("/api/vehicles/excel/template", headers=ADMIN_HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["车辆信息"]
    sheet.append([None, None, "品牌A", None, None, "不是数字", "2021-99", -1])
    content = BytesIO()
    workbook.save(content)
    preview = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("bad.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["error_count"] == 1
    assert data["can_confirm"] is False
    confirm = client.post(f"/api/vehicles/excel/{data['preview_id']}/confirm", headers=ADMIN_HEADERS)
    assert confirm.status_code == 409
    assert confirm.json()["code"] == "VEHICLE_IMPORT_HAS_ERRORS"
    with SessionLocal() as db:
        assert db.scalar(select(KnowledgeItem)) is None
        stored_preview = db.get(VehicleImportPreview, data["preview_id"])
        assert stored_preview.status == "pending"


def test_excel_nonexistent_vehicle_code_is_an_error_not_a_create():
    template = client.get("/api/vehicles/excel/template", headers=ADMIN_HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["车辆信息"]
    sheet.append(["CJ-NOT-EXISTS", "不能被创建", "品牌A", "车系A", None, 9.98, "2021-08", 31000])
    content = BytesIO()
    workbook.save(content)

    preview = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("vehicles.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["create_count"] == 0
    assert data["update_count"] == 0
    assert data["error_count"] == 1
    assert data["can_confirm"] is False
    assert "车辆编号不存在" in data["rows"][0]["errors"][0]
    with SessionLocal() as db:
        assert db.scalar(select(KnowledgeItem).where(KnowledgeItem.item_id == "CJ-NOT-EXISTS")) is None


def test_vehicle_list_search_filter_detail_update_and_system_fields_are_protected():
    first = _create_vehicle(display_name="星河运动轿车", brand="星河", series="闪电")
    second = _create_vehicle(display_name="远山城市车", brand="远山", series="通勤")

    by_name = client.get("/api/vehicles?keyword=星河", headers=ADMIN_HEADERS).json()["data"]
    by_code = client.get(f"/api/vehicles?keyword={second['vehicle_code']}", headers=ADMIN_HEADERS).json()["data"]
    by_series = client.get("/api/vehicles?keyword=通勤", headers=ADMIN_HEADERS).json()["data"]
    assert [item["vehicle_code"] for item in by_name["items"]] == [first["vehicle_code"]]
    assert [item["vehicle_code"] for item in by_code["items"]] == [second["vehicle_code"]]
    assert [item["vehicle_code"] for item in by_series["items"]] == [second["vehicle_code"]]

    all_rows = client.get("/api/vehicles?listing_status=all", headers=ADMIN_HEADERS).json()["data"]
    unlisted = client.get("/api/vehicles?listing_status=unlisted", headers=ADMIN_HEADERS).json()["data"]
    listed = client.get("/api/vehicles?listing_status=listed", headers=ADMIN_HEADERS).json()["data"]
    assert (all_rows["total"], unlisted["total"], listed["total"]) == (2, 2, 0)

    updated = client.put(
        f"/api/vehicles/{first['vehicle_code']}",
        json={"brand": "星河汽车", "public_price": 16.88},
        headers=ADMIN_HEADERS,
    )
    assert updated.status_code == 200
    detail = client.get(f"/api/vehicles/{first['vehicle_code']}", headers=ADMIN_HEADERS).json()["data"]
    assert detail["brand"] == "星河汽车"
    assert detail["vehicle_code"] == first["vehicle_code"]
    assert detail["created_at"] and detail["updated_at"]

    protected = client.put(
        f"/api/vehicles/{first['vehicle_code']}",
        json={"vehicle_code": "FORGED", "created_at": "2020-01-01", "updated_at": "2020-01-01"},
        headers=ADMIN_HEADERS,
    )
    assert protected.status_code == 400
    unchanged = client.get(f"/api/vehicles/{first['vehicle_code']}", headers=ADMIN_HEADERS).json()["data"]
    assert unchanged["vehicle_code"] == first["vehicle_code"]
    assert unchanged["created_at"] == detail["created_at"]
    with SessionLocal() as db:
        assert db.scalar(select(OperationLog).where(OperationLog.event_type == "vehicle_updated")) is not None


def test_vehicle_image_formats_order_read_delete_and_audit_log_are_real():
    vehicle = _create_vehicle(public_price=12.88)
    code = vehicle["vehicle_code"]
    uploaded = client.post(
        f"/api/vehicles/{code}/images",
        files=[
            ("files", ("one.jpg", JPEG_MINIMAL, "image/jpeg")),
            ("files", ("two.png", PNG_1X1, "image/png")),
            ("files", ("three.webp", WEBP_MINIMAL, "image/webp")),
            ("files", ("bad.gif", b"GIF89a", "image/gif")),
            ("files", ("too-large.jpg", b"\xff\xd8\xff" + b"x" * (10 * 1024 * 1024) + b"\xff\xd9", "image/jpeg")),
        ],
        headers=ADMIN_HEADERS,
    )
    assert uploaded.status_code == 200
    data = uploaded.json()["data"]
    assert data["succeeded"] == 3
    assert data["failed"] == 2
    failures = {item["filename"]: item["error_code"] for item in data["items"] if not item["ok"]}
    assert failures == {"bad.gif": "VEHICLE_IMAGE_DECODE_FAILED", "too-large.jpg": "VEHICLE_IMAGE_TOO_LARGE"}
    image_ids = [item["image"]["id"] for item in data["items"] if item["ok"]]

    for image_id, expected_type in zip(image_ids, ("image/jpeg", "image/png", "image/webp")):
        response = client.get(f"/api/vehicles/images/{image_id}", headers=ADMIN_HEADERS)
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(expected_type)

    reversed_ids = list(reversed(image_ids))
    reordered = client.put(
        f"/api/vehicles/{code}/images/order",
        json={"image_ids": reversed_ids},
        headers=ADMIN_HEADERS,
    )
    assert reordered.status_code == 200
    assert [item["id"] for item in reordered.json()["data"]["items"]] == reversed_ids
    assert reordered.json()["data"]["items"][0]["is_main"] is True

    assert client.delete(f"/api/vehicles/{code}/images/{reversed_ids[1]}", headers=ADMIN_HEADERS).status_code == 200
    assert client.post(f"/api/vehicles/{code}/list", headers=ADMIN_HEADERS).status_code == 200
    assert client.post(f"/api/vehicles/{code}/unlist", headers=ADMIN_HEADERS).status_code == 200

    with SessionLocal() as db:
        event_types = set(db.scalars(select(OperationLog.event_type).where(OperationLog.module == "vehicles")).all())
    assert {
        "vehicle_created",
        "vehicle_image_uploaded",
        "vehicle_image_reordered",
        "vehicle_image_deleted",
        "vehicle_listed",
        "vehicle_unlisted",
    } <= event_types


def test_vehicle_image_upload_enforces_batch_limits_and_real_decode(monkeypatch):
    vehicle_id = _create_vehicle()["vehicle_code"]
    settings = get_settings()
    monkeypatch.setattr(settings, "vehicle_image_upload_max_files", 2)
    too_many = client.post(
        f"/api/vehicles/{vehicle_id}/images",
        files=[
            ("files", (f"{index}.png", _image_bytes("PNG", (index, 0, 0)), "image/png"))
            for index in range(3)
        ],
        headers=ADMIN_HEADERS,
    )
    assert too_many.status_code == 413
    assert too_many.json()["code"] == "VEHICLE_IMAGE_UPLOAD_FILE_COUNT_EXCEEDED"

    monkeypatch.setattr(settings, "vehicle_image_upload_max_files", 20)
    monkeypatch.setattr(settings, "vehicle_image_upload_max_total_bytes", len(PNG_1X1) + 1)
    too_large_batch = client.post(
        f"/api/vehicles/{vehicle_id}/images",
        files=[
            ("files", ("one.png", PNG_1X1, "image/png")),
            ("files", ("two.png", _image_bytes("PNG", (1, 2, 3)), "image/png")),
        ],
        headers=ADMIN_HEADERS,
    )
    assert too_large_batch.status_code == 413
    assert too_large_batch.json()["code"] == "VEHICLE_IMAGE_UPLOAD_TOTAL_TOO_LARGE"

    monkeypatch.setattr(settings, "vehicle_image_upload_max_total_bytes", 30 * 1024 * 1024)
    malformed = client.post(
        f"/api/vehicles/{vehicle_id}/images",
        files=[("files", ("forged.jpg", b"\xff\xd8\xffnot-a-real-image\xff\xd9", "image/jpeg"))],
        headers=ADMIN_HEADERS,
    )
    assert malformed.status_code == 200
    assert malformed.json()["data"]["items"][0]["error_code"] == "VEHICLE_IMAGE_DECODE_FAILED"

    with SessionLocal() as db:
        failures = list(db.scalars(select(OperationLog).where(OperationLog.event_type == "vehicle_operation_failed")))
        assert any((item.extra_metadata or {}).get("operation") == "upload_images" for item in failures)
        assert not list(db.scalars(select(VehicleImage).where(VehicleImage.vehicle_id == vehicle_id)))


def test_vehicle_failure_audit_survives_business_rollback():
    vehicle_id = _create_vehicle()["vehicle_code"]
    failed = client.post(f"/api/vehicles/{vehicle_id}/list", headers=ADMIN_HEADERS)
    assert failed.status_code == 409
    assert failed.json()["code"] == "VEHICLE_LISTING_PRICE_REQUIRED"
    with SessionLocal() as db:
        log = db.scalar(
            select(OperationLog)
            .where(OperationLog.event_type == "vehicle_operation_failed")
            .order_by(OperationLog.created_at.desc())
        )
        assert log is not None
        assert log.target_id == vehicle_id
        assert log.request_id
        assert log.extra_metadata == {
            "operation": "list",
            "error_code": "VEHICLE_LISTING_PRICE_REQUIRED",
            "error_type": "AppError",
            "status_code": 409,
        }


def test_vehicle_image_delete_tracks_disk_cleanup_until_success(monkeypatch):
    vehicle_id = _create_vehicle()["vehicle_code"]
    uploaded = client.post(
        f"/api/vehicles/{vehicle_id}/images",
        files=[("files", ("one.png", PNG_1X1, "image/png"))],
        headers=ADMIN_HEADERS,
    ).json()["data"]["items"][0]["image"]
    path = vehicle_service._storage_path(f"{get_settings().omniauto_knowledge_tenant}/{vehicle_id}/{uploaded['id']}.png")
    assert path.is_file()

    original_unlink = Path.unlink
    monkeypatch.setattr(Path, "unlink", lambda self, missing_ok=False: (_ for _ in ()).throw(OSError("disk busy")))
    deleted = client.delete(f"/api/vehicles/{vehicle_id}/images/{uploaded['id']}", headers=ADMIN_HEADERS)
    assert deleted.status_code == 200
    assert deleted.json()["data"]["file_cleanup_pending"] is True

    with SessionLocal() as db:
        assert db.get(VehicleImage, uploaded["id"]) is None
        cleanup = db.scalar(select(VehicleFileCleanup).where(VehicleFileCleanup.storage_key.endswith(f"/{uploaded['id']}.png")))
        assert cleanup is not None
        assert cleanup.status == "pending"
        assert cleanup.attempts == 1
        cleanup_id = cleanup.id
    monkeypatch.setattr(Path, "unlink", original_unlink)
    assert vehicle_service.process_vehicle_file_cleanup(cleanup_id) is True
    assert not path.exists()
    with SessionLocal() as db:
        assert db.get(VehicleFileCleanup, cleanup_id).status == "completed"


def test_excel_update_preserves_blank_fields_and_confirm_rolls_back_the_whole_batch():
    existing = _create_vehicle(display_name="原展示名", brand="原品牌", series="原车系", public_price=20.0)
    template = client.get("/api/vehicles/excel/template", headers=ADMIN_HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["车辆信息"]
    sheet.append([existing["vehicle_code"], None, "新品牌"])
    content = BytesIO()
    workbook.save(content)
    preview = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("update.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    ).json()["data"]
    assert preview["update_count"] == 1 and preview["can_confirm"] is True
    assert client.post(f"/api/vehicles/excel/{preview['preview_id']}/confirm", headers=ADMIN_HEADERS).status_code == 200
    updated = client.get(f"/api/vehicles/{existing['vehicle_code']}", headers=ADMIN_HEADERS).json()["data"]
    assert updated["display_name"] == "原展示名"
    assert updated["brand"] == "新品牌"
    assert updated["series"] == "原车系"
    assert float(updated["public_price"]) == 20.0
    with SessionLocal() as db:
        assert db.scalar(select(OperationLog).where(OperationLog.event_type == "vehicle_excel_import_confirmed")) is not None

    rollback_book = load_workbook(BytesIO(template.content))
    rollback_sheet = rollback_book["车辆信息"]
    rollback_sheet.append([None, "整批第一辆"])
    rollback_sheet.append([None, "整批第二辆"])
    rollback_content = BytesIO()
    rollback_book.save(rollback_content)
    rollback_preview = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("rollback.xlsx", rollback_content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    ).json()["data"]
    first_code = rollback_preview["rows"][0]["vehicle_code"]
    second_code = rollback_preview["rows"][1]["vehicle_code"]

    actor = ActorContext(
        operator_id=UUID("00000000-0000-0000-0000-000000000001"),
        operator_name="Concurrent Writer",
        role="admin",
        ip_address=None,
        user_agent=None,
        request_id="vehicle-rollback-acceptance",
    )
    with SessionLocal() as db:
        vehicle_service._create_with_id(db, second_code, {"display_name": "并发占用车辆"}, actor)
        db.commit()

    failed = client.post(f"/api/vehicles/excel/{rollback_preview['preview_id']}/confirm", headers=ADMIN_HEADERS)
    assert failed.status_code == 409
    assert failed.json()["code"] == "VEHICLE_IMPORT_CONCURRENT_CONFLICT"
    with SessionLocal() as db:
        assert db.scalar(select(KnowledgeItem).where(KnowledgeItem.item_id == first_code)) is None
        assert db.scalar(select(KnowledgeItem).where(KnowledgeItem.item_id == second_code)) is not None


def test_excel_template_file_size_type_and_row_limits_match_the_acceptance_contract():
    template = client.get("/api/vehicles/excel/template", headers=ADMIN_HEADERS)
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    assert {"车辆信息", "字段说明", "_元数据"} <= set(workbook.sheetnames)
    descriptions = {row[0].value: row[1].value for row in workbook["字段说明"].iter_rows(min_row=2) if row[0].value}
    assert "新增必须留空" in descriptions["车辆编号"]
    assert "不存在编号会报错" in descriptions["车辆编号"]

    wrong_type = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("vehicles.xls", b"legacy", "application/vnd.ms-excel")},
        headers=ADMIN_HEADERS,
    )
    assert wrong_type.status_code == 400
    assert wrong_type.json()["code"] == "VEHICLE_EXCEL_TYPE_INVALID"

    too_large = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("vehicles.xlsx", b"x" * (8 * 1024 * 1024 + 1), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    )
    assert too_large.status_code == 413
    assert too_large.json()["code"] == "VEHICLE_EXCEL_TOO_LARGE"

    row_book = load_workbook(BytesIO(template.content))
    row_sheet = row_book["车辆信息"]
    for index in range(2001):
        row_sheet.append([None, f"第 {index + 1} 辆"])
    row_content = BytesIO()
    row_book.save(row_content)
    too_many_rows = client.post(
        "/api/vehicles/excel/preview",
        files={"file": ("vehicles.xlsx", row_content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=ADMIN_HEADERS,
    )
    assert too_many_rows.status_code == 413
    assert too_many_rows.json()["code"] == "VEHICLE_EXCEL_ROWS_EXCEEDED"
